"""Synthetic-only tests for the u139 private XLSX input boundary."""

from __future__ import annotations

import ast
import json
from collections.abc import Callable, Mapping, Sequence
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import pytest
from openpyxl import Workbook  # type: ignore[import-untyped]

import investo.sector_dashboard.private_input as subject
from investo.models import (
    ALL_SECTOR_TICKERS,
    DiagnosticIssueCode,
    NavSeries,
    ParsedWorkbookSet,
    PrivateWorkbookManifest,
    SectorCoverageStatus,
    SectorTicker,
    WorkbookFailure,
    WorkbookIssueCode,
)
from investo.sector_dashboard import (
    PrivateInputError,
    load_private_nav_workbooks,
    parse_private_nav_workbooks,
    read_private_workbook_manifest,
)

REPOSITORY_ROOT = Path(__file__).resolve().parents[3]
_START = date(2026, 1, 2)


def _rows(days: int) -> list[tuple[object, object]]:
    return [(_START + timedelta(days=index), 100 + index) for index in range(days)]


def _write_workbook(
    path: Path,
    rows: Sequence[tuple[object, object]],
    *,
    header: tuple[str, str] = (" Date ", "nav"),
    duplicate_header_sheet: bool = False,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["synthetic metadata"])
    worksheet.append([header[0], header[1], "Ignored Synthetic Column"])
    for trading_date, nav in rows:
        worksheet.append([trading_date, nav, "ignored"])
    if duplicate_header_sheet:
        other = workbook.create_sheet("synthetic-other")
        other.append(["DATE", "NAV"])
        other.append([_START, 100])
        other.append([_START + timedelta(days=1), 101])
    workbook.save(path)
    workbook.close()


def _write_manifest_files(
    tmp_path: Path,
    *,
    days: int = 22,
    rows_by_ticker: Mapping[SectorTicker, Sequence[tuple[object, object]]] | None = None,
    header_by_ticker: Mapping[SectorTicker, tuple[str, str]] | None = None,
    duplicate_header_tickers: frozenset[SectorTicker] = frozenset(),
) -> PrivateWorkbookManifest:
    overrides = rows_by_ticker or {}
    headers = header_by_ticker or {}
    paths: dict[SectorTicker, Path] = {}
    for ticker in ALL_SECTOR_TICKERS:
        path = tmp_path / f"{ticker.value.lower()}.xlsx"
        _write_workbook(
            path,
            overrides.get(ticker, _rows(days)),
            header=headers.get(ticker, (" Date ", "nav")),
            duplicate_header_sheet=ticker in duplicate_header_tickers,
        )
        paths[ticker] = path
    return PrivateWorkbookManifest(schema_version=1, workbooks=paths)


def _failure_for(
    parsed: ParsedWorkbookSet,
    ticker: SectorTicker,
) -> WorkbookFailure:
    return next(failure for failure in parsed.failures if failure.ticker is ticker)


def _rewrite_zip_member(path: Path, member_name: str, transform: Callable[[bytes], bytes]) -> None:
    replacement = path.with_suffix(".replacement")
    with ZipFile(path) as source, ZipFile(replacement, "w", ZIP_DEFLATED) as target:
        for info in source.infolist():
            payload = source.read(info.filename)
            target.writestr(info, transform(payload) if info.filename == member_name else payload)
    replacement.replace(path)


def test_read_manifest_accepts_only_the_fixed_explicit_mapping(tmp_path: Path) -> None:
    manifest = _write_manifest_files(tmp_path)
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "workbooks": {
                    ticker.value: str(manifest.workbooks[ticker])
                    for ticker in reversed(ALL_SECTOR_TICKERS)
                },
            }
        ),
        encoding="utf-8",
    )

    loaded = read_private_workbook_manifest(
        manifest_path,
        repository_root=REPOSITORY_ROOT,
    )

    assert tuple(loaded.workbooks) == ALL_SECTOR_TICKERS
    assert all(path.is_absolute() for path in loaded.workbooks.values())


def test_manifest_schema_and_universe_failures_are_closed_and_redacted(tmp_path: Path) -> None:
    for name, payload, expected in (
        (
            "schema.json",
            {"schema_version": 2, "workbooks": {}},
            DiagnosticIssueCode.MANIFEST_SCHEMA,
        ),
        (
            "universe.json",
            {"schema_version": 1, "workbooks": {"SPY": "/private/spy.xlsx"}},
            DiagnosticIssueCode.MANIFEST_UNIVERSE,
        ),
        (
            "extra-key.json",
            {"schema_version": 1, "workbooks": {}, "private_path": "/sentinel"},
            DiagnosticIssueCode.MANIFEST_SCHEMA,
        ),
    ):
        path = tmp_path / name
        path.write_text(json.dumps(payload), encoding="utf-8")

        with pytest.raises(PrivateInputError) as caught:
            read_private_workbook_manifest(path, repository_root=REPOSITORY_ROOT)

        assert caught.value.issue_code is expected
        assert str(caught.value) == expected.value
        assert caught.value.__cause__ is None


def test_manifest_read_is_bounded_by_the_schema_ceiling(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    manifest_path = tmp_path / "oversized.json"
    manifest_path.write_bytes(b"{" + b" " * 32 + b"}")
    monkeypatch.setattr(subject, "MAX_MANIFEST_BYTES", 16)

    with pytest.raises(PrivateInputError) as caught:
        read_private_workbook_manifest(manifest_path, repository_root=REPOSITORY_ROOT)

    assert caught.value.issue_code is DiagnosticIssueCode.MANIFEST_SCHEMA
    assert caught.value.__cause__ is None


def test_manifest_rejects_duplicate_json_identity_keys(tmp_path: Path) -> None:
    manifest = _write_manifest_files(tmp_path)
    entries = [
        f"{json.dumps(ticker.value)}:{json.dumps(str(manifest.workbooks[ticker]))}"
        for ticker in ALL_SECTOR_TICKERS
    ]
    entries.insert(1, entries[0])
    manifest_path = tmp_path / "duplicate-key.json"
    manifest_path.write_text(
        '{"schema_version":1,"workbooks":{' + ",".join(entries) + "}}",
        encoding="utf-8",
    )

    with pytest.raises(PrivateInputError) as caught:
        read_private_workbook_manifest(manifest_path, repository_root=REPOSITORY_ROOT)

    assert caught.value.issue_code is DiagnosticIssueCode.MANIFEST_SCHEMA
    assert caught.value.__cause__ is None


def test_manifest_replacement_after_identity_check_fails_closed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    manifest = _write_manifest_files(tmp_path)
    manifest_path = tmp_path / "manifest.json"
    replacement = tmp_path / "replacement.json"
    payload = {
        "schema_version": 1,
        "workbooks": {
            ticker.value: str(manifest.workbooks[ticker]) for ticker in ALL_SECTOR_TICKERS
        },
    }
    manifest_path.write_text(json.dumps(payload), encoding="utf-8")
    replacement.write_text(json.dumps(payload), encoding="utf-8")
    real_resolve = subject._resolve_external_regular_file
    swapped = False

    def swapping_resolve(
        path: Path,
        *,
        repository_root: Path,
        expected_suffix: str,
    ) -> tuple[Path, subject._FileIdentity]:
        nonlocal swapped
        result = real_resolve(
            path,
            repository_root=repository_root,
            expected_suffix=expected_suffix,
        )
        if not swapped:
            manifest_path.unlink()
            replacement.replace(manifest_path)
            swapped = True
        return result

    monkeypatch.setattr(subject, "_resolve_external_regular_file", swapping_resolve)
    with pytest.raises(PrivateInputError) as caught:
        read_private_workbook_manifest(manifest_path, repository_root=REPOSITORY_ROOT)

    assert swapped
    assert caught.value.issue_code is DiagnosticIssueCode.MANIFEST_PATH
    assert caught.value.__cause__ is None


def test_path_rejection_is_terminal_redacted_and_precedes_workbook_open(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    manifest = _write_manifest_files(tmp_path)
    sentinel = tmp_path / "private-operator-alias.xlsx"
    sentinel.symlink_to(manifest.workbooks[SectorTicker.XLC])
    unsafe = PrivateWorkbookManifest(
        schema_version=1,
        workbooks={**manifest.workbooks, SectorTicker.XLC: sentinel},
    )

    def forbidden_open(*args: object, **kwargs: object) -> object:
        raise AssertionError("openpyxl must not run before every path is valid")

    monkeypatch.setattr(subject, "_load_workbook", forbidden_open)
    with pytest.raises(PrivateInputError) as caught:
        load_private_nav_workbooks(unsafe, repository_root=REPOSITORY_ROOT)

    assert caught.value.issue_code is DiagnosticIssueCode.MANIFEST_PATH
    assert str(caught.value) == "manifest.path"
    assert str(sentinel) not in str(caught.value)
    assert caught.value.__cause__ is None


def test_hardlink_aliases_are_rejected_as_duplicate_private_inputs(tmp_path: Path) -> None:
    manifest = _write_manifest_files(tmp_path)
    alias = tmp_path / "xlc-hardlink.xlsx"
    alias.hardlink_to(manifest.workbooks[SectorTicker.XLC])
    unsafe = PrivateWorkbookManifest(
        schema_version=1,
        workbooks={**manifest.workbooks, SectorTicker.XLY: alias},
    )

    with pytest.raises(PrivateInputError) as caught:
        parse_private_nav_workbooks(unsafe, repository_root=REPOSITORY_ROOT)

    assert caught.value.issue_code is DiagnosticIssueCode.MANIFEST_PATH
    assert caught.value.__cause__ is None


def test_load_normalizes_descending_series_and_builds_normal_bundle(tmp_path: Path) -> None:
    manifest = _write_manifest_files(
        tmp_path,
        rows_by_ticker={SectorTicker.XLC: tuple(reversed(_rows(22)))},
    )

    bundle = load_private_nav_workbooks(manifest, repository_root=REPOSITORY_ROOT)

    assert bundle.coverage.status is SectorCoverageStatus.NORMAL
    assert bundle.coverage.available_sector_count == 11
    assert bundle.coverage.benchmark_observation_count == 22
    assert bundle.as_of_date == _START + timedelta(days=21)
    assert bundle.input_fingerprint is not None
    assert bundle.input_fingerprint.startswith("sha256:")
    xlc = next(series for series in bundle.sectors if series.ticker is SectorTicker.XLC)
    assert xlc.first_date == _START
    assert xlc.latest_date == bundle.as_of_date
    assert tuple(point.trading_date for point in xlc.points) == tuple(
        sorted(point.trading_date for point in xlc.points)
    )
    assert bundle.diagnostics == ()


def test_input_fingerprint_is_stable_and_changes_with_one_workbook(tmp_path: Path) -> None:
    manifest = _write_manifest_files(tmp_path)

    first = load_private_nav_workbooks(manifest, repository_root=REPOSITORY_ROOT)
    repeated = load_private_nav_workbooks(manifest, repository_root=REPOSITORY_ROOT)
    changed_rows = _rows(22)
    changed_rows[3] = (changed_rows[3][0], 500)
    _write_workbook(manifest.workbooks[SectorTicker.XLC], changed_rows)
    changed = load_private_nav_workbooks(manifest, repository_root=REPOSITORY_ROOT)

    assert first.input_fingerprint == repeated.input_fingerprint
    assert changed.input_fingerprint != first.input_fingerprint


def test_one_invalid_sector_is_isolated_and_yields_partial_redacted_bundle(
    tmp_path: Path,
) -> None:
    invalid_rows = _rows(22)
    invalid_rows[8] = (invalid_rows[8][0], 0)
    manifest = _write_manifest_files(
        tmp_path,
        rows_by_ticker={SectorTicker.XLC: invalid_rows},
    )

    bundle = load_private_nav_workbooks(manifest, repository_root=REPOSITORY_ROOT)

    assert bundle.coverage.status is SectorCoverageStatus.PARTIAL
    assert bundle.coverage.available_sector_count == 10
    assert bundle.coverage.missing_tickers == (SectorTicker.XLC,)
    assert bundle.input_fingerprint is not None
    diagnostic = next(item for item in bundle.diagnostics if item.ticker is SectorTicker.XLC)
    assert diagnostic.issue_code is DiagnosticIssueCode.WORKBOOK_NAV
    assert str(tmp_path) not in bundle.model_dump_json()


@pytest.mark.parametrize(
    ("case", "expected_issue"),
    [
        ("missing-header", WorkbookIssueCode.HEADER),
        ("duplicate-header", WorkbookIssueCode.HEADER),
        ("duplicate-date", WorkbookIssueCode.DUPLICATE_DATE),
        ("interior-order", WorkbookIssueCode.ORDER),
        ("timestamp", WorkbookIssueCode.DATE),
        ("formula-without-value", WorkbookIssueCode.NAV),
    ],
)
def test_parser_rejects_malformed_synthetic_workbooks_with_typed_failures(
    tmp_path: Path,
    case: str,
    expected_issue: WorkbookIssueCode,
) -> None:
    rows = _rows(22)
    headers: dict[SectorTicker, tuple[str, str]] = {}
    duplicate_headers: frozenset[SectorTicker] = frozenset()
    if case == "missing-header":
        headers[SectorTicker.XLC] = ("When", "Value")
    elif case == "duplicate-header":
        duplicate_headers = frozenset({SectorTicker.XLC})
    elif case == "duplicate-date":
        rows[6] = (rows[5][0], rows[6][1])
    elif case == "interior-order":
        rows[5], rows[6] = rows[6], rows[5]
    elif case == "timestamp":
        rows[5] = (datetime(2026, 1, 7, 12, 30), rows[5][1])
    elif case == "formula-without-value":
        rows[5] = (rows[5][0], "=100+5")
    manifest = _write_manifest_files(
        tmp_path,
        rows_by_ticker={SectorTicker.XLC: rows},
        header_by_ticker=headers,
        duplicate_header_tickers=duplicate_headers,
    )

    parsed = parse_private_nav_workbooks(manifest, repository_root=REPOSITORY_ROOT)
    failure = _failure_for(parsed, SectorTicker.XLC)

    assert failure.issue_code is expected_issue
    assert not hasattr(failure, "path")
    assert not hasattr(failure, "message")


def test_as_of_mismatch_suppresses_bundle_comparability(tmp_path: Path) -> None:
    mismatched = _rows(22)
    mismatched[-1] = (_START + timedelta(days=22), mismatched[-1][1])
    manifest = _write_manifest_files(
        tmp_path,
        rows_by_ticker={SectorTicker.XLK: mismatched},
    )

    bundle = load_private_nav_workbooks(manifest, repository_root=REPOSITORY_ROOT)

    assert bundle.coverage.status is SectorCoverageStatus.INSUFFICIENT
    assert bundle.as_of_date is None
    assert DiagnosticIssueCode.BUNDLE_AS_OF_MISMATCH in bundle.coverage.reason_codes


def test_spy_failure_is_a_bundle_level_insufficient_result(tmp_path: Path) -> None:
    invalid_spy_rows = _rows(22)
    invalid_spy_rows[4] = (invalid_spy_rows[4][0], 0)
    manifest = _write_manifest_files(
        tmp_path,
        rows_by_ticker={SectorTicker.SPY: invalid_spy_rows},
    )

    bundle = load_private_nav_workbooks(manifest, repository_root=REPOSITORY_ROOT)

    assert bundle.coverage.status is SectorCoverageStatus.INSUFFICIENT
    assert bundle.benchmark is None
    assert bundle.as_of_date is None
    assert bundle.coverage.benchmark_observation_count == 0
    assert DiagnosticIssueCode.BUNDLE_SPY_MISSING in bundle.coverage.reason_codes


@pytest.mark.parametrize(
    ("days", "expected"),
    [
        (6, SectorCoverageStatus.WARMING_UP),
        (5, SectorCoverageStatus.INSUFFICIENT),
    ],
)
def test_observation_count_resolves_warming_and_insufficient_states(
    tmp_path: Path,
    days: int,
    expected: SectorCoverageStatus,
) -> None:
    manifest = _write_manifest_files(tmp_path, days=days)

    bundle = load_private_nav_workbooks(manifest, repository_root=REPOSITORY_ROOT)

    assert bundle.coverage.status is expected
    assert bundle.coverage.benchmark_observation_count == days


@pytest.mark.parametrize("unsafe_name", ["../escape.xml", "/absolute.xml", "C:/drive.xml"])
def test_zip_metadata_rejects_unsafe_member_paths(unsafe_name: str) -> None:
    info = ZipInfo(unsafe_name)
    info.file_size = 1
    info.compress_size = 1

    with pytest.raises(subject._WorkbookRejectedError):
        subject._validate_zip_metadata((info,), compressed_size=1)


def test_zip_metadata_rejects_duplicates_encryption_and_excessive_ratio() -> None:
    first = ZipInfo("safe.xml")
    first.file_size = 1
    first.compress_size = 1
    duplicate = ZipInfo("safe.xml")
    duplicate.file_size = 1
    duplicate.compress_size = 1
    encrypted = ZipInfo("encrypted.xml")
    encrypted.file_size = 1
    encrypted.compress_size = 1
    encrypted.flag_bits = 0x1
    compressed = ZipInfo("compressed.xml")
    compressed.file_size = 101
    compressed.compress_size = 1

    for infos in ((first, duplicate), (encrypted,), (compressed,)):
        with pytest.raises(subject._WorkbookRejectedError):
            subject._validate_zip_metadata(infos, compressed_size=1)


def test_zip_metadata_rejects_member_and_uncompressed_size_ceilings(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    info = ZipInfo("safe.xml")
    info.file_size = 2
    info.compress_size = 2

    monkeypatch.setattr(subject, "MAX_UNCOMPRESSED_WORKBOOK_BYTES", 1)
    with pytest.raises(subject._WorkbookRejectedError):
        subject._validate_zip_metadata((info,), compressed_size=2)
    monkeypatch.setattr(subject, "MAX_UNCOMPRESSED_WORKBOOK_BYTES", 64 * 1024 * 1024)
    monkeypatch.setattr(subject, "MAX_ZIP_MEMBERS", 0)
    with pytest.raises(subject._WorkbookRejectedError):
        subject._validate_zip_metadata((info,), compressed_size=2)


def test_preflight_rejection_never_reaches_openpyxl(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    manifest = _write_manifest_files(tmp_path, days=2)

    def forbidden_open(*args: object, **kwargs: object) -> object:
        raise AssertionError("rejected packages must not reach openpyxl")

    monkeypatch.setattr(subject, "MAX_COMPRESSED_WORKBOOK_BYTES", 1)
    monkeypatch.setattr(subject, "_load_workbook", forbidden_open)
    parsed = parse_private_nav_workbooks(manifest, repository_root=REPOSITORY_ROOT)

    assert len(parsed.failures) == 12
    assert {failure.issue_code for failure in parsed.failures} == {WorkbookIssueCode.OPEN}


def test_preflight_rejects_malformed_non_xlsx_and_defused_xml(tmp_path: Path) -> None:
    malformed = tmp_path / "malformed.xlsx"
    malformed.write_bytes(b"not-a-zip")
    wrong_type = tmp_path / "wrong-type.xlsx"
    entity_xml = tmp_path / "entity.xlsx"
    safe_workbook_xml = (
        b'<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheets><sheet name="Synthetic" sheetId="1"/></sheets></workbook>'
    )
    safe_sheet_xml = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetData><row r="1"><c r="A1"/></row></sheetData></worksheet>'
    )
    rels_xml = (
        b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>'
    )
    wrong_content_types = (
        b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        b'<Override PartName="/xl/workbook.xml" ContentType="application/msword"/></Types>'
    )
    entity_content_types = (
        b'<!DOCTYPE Types [<!ENTITY private "private">]>'
        b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        b'<Override PartName="&private;" ContentType="application/msword"/></Types>'
    )
    for path, content_types in (
        (wrong_type, wrong_content_types),
        (entity_xml, entity_content_types),
    ):
        with ZipFile(path, "w", ZIP_DEFLATED) as package:
            package.writestr("[Content_Types].xml", content_types)
            package.writestr("_rels/.rels", rels_xml)
            package.writestr("xl/workbook.xml", safe_workbook_xml)
            package.writestr("xl/worksheets/sheet1.xml", safe_sheet_xml)

    for path in (malformed, wrong_type, entity_xml):
        with pytest.raises(subject._WorkbookRejectedError) as caught:
            subject._preflight_xlsx(path)
        assert caught.value.issue_code is WorkbookIssueCode.OPEN


def test_shape_preflight_counts_cell_records_and_worksheets(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook_path = tmp_path / "bounded.xlsx"
    _write_workbook(workbook_path, _rows(2))

    monkeypatch.setattr(subject, "MAX_WORKSHEET_CELLS", 1)
    with pytest.raises(subject._WorkbookRejectedError):
        subject._preflight_xlsx(workbook_path)
    monkeypatch.setattr(subject, "MAX_WORKSHEET_CELLS", 250_000)
    monkeypatch.setattr(subject, "MAX_WORKSHEETS", 0)
    with pytest.raises(subject._WorkbookRejectedError):
        subject._preflight_xlsx(workbook_path)


def test_shape_preflight_counts_the_actual_relationship_target(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook_path = tmp_path / "relationship-target.xlsx"
    replacement = tmp_path / "relationship-target.replacement"
    _write_workbook(workbook_path, _rows(2))
    benign_sheet = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetData><row r="1"><c r="A1"/></row></sheetData></worksheet>'
    )
    with ZipFile(workbook_path) as source, ZipFile(replacement, "w", ZIP_DEFLATED) as target:
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                target.writestr(info, benign_sheet)
                target.writestr("xl/custom/actual-sheet.xml", payload)
            elif info.filename == "xl/_rels/workbook.xml.rels":
                rewritten = payload.replace(
                    b'Target="/xl/worksheets/sheet1.xml"',
                    b'Target="/xl/custom/actual-sheet.xml"',
                )
                assert rewritten != payload
                target.writestr(info, rewritten)
            elif info.filename == "[Content_Types].xml":
                rewritten = payload.replace(
                    b'PartName="/xl/worksheets/sheet1.xml"',
                    b'PartName="/xl/custom/actual-sheet.xml"',
                )
                assert rewritten != payload
                target.writestr(info, rewritten)
            else:
                target.writestr(info, payload)
    replacement.replace(workbook_path)

    monkeypatch.setattr(subject, "MAX_WORKSHEET_CELLS", 1)
    with pytest.raises(subject._WorkbookRejectedError):
        subject._preflight_xlsx(workbook_path)


@pytest.mark.parametrize("unsafe_target", ["../../escape.xml", "external"])
def test_preflight_rejects_unsafe_or_external_worksheet_relationships(
    tmp_path: Path,
    unsafe_target: str,
) -> None:
    workbook_path = tmp_path / f"relationship-{unsafe_target.replace('/', '-')}.xlsx"
    _write_workbook(workbook_path, _rows(2))

    def rewrite_relationship(payload: bytes) -> bytes:
        original = b'Target="/xl/worksheets/sheet1.xml"'
        replacement = (
            b'Target="/xl/worksheets/sheet1.xml" TargetMode="External"'
            if unsafe_target == "external"
            else f'Target="{unsafe_target}"'.encode()
        )
        rewritten = payload.replace(original, replacement)
        assert rewritten != payload
        return rewritten

    _rewrite_zip_member(
        workbook_path,
        "xl/_rels/workbook.xml.rels",
        rewrite_relationship,
    )

    with pytest.raises(subject._WorkbookRejectedError):
        subject._preflight_xlsx(workbook_path)


def test_misleading_worksheet_dimension_is_ignored_before_streaming(tmp_path: Path) -> None:
    workbook_path = tmp_path / "misleading-dimension.xlsx"
    _write_workbook(workbook_path, _rows(2))

    def exaggerate_dimension(payload: bytes) -> bytes:
        return payload.replace(b'ref="A1:C4"', b'ref="A1:XFD1048576"')

    _rewrite_zip_member(workbook_path, "xl/worksheets/sheet1.xml", exaggerate_dimension)
    subject._preflight_xlsx(workbook_path)
    resolved = subject._ResolvedWorkbook(
        path=workbook_path,
        identity=subject._FileIdentity.from_stat(workbook_path.stat()),
    )
    result, workbook_digest = subject._parse_verified_workbook(SectorTicker.XLC, resolved)

    assert isinstance(result, NavSeries)
    assert len(result.points) == 2
    assert workbook_digest is not None


def test_path_replacement_during_parse_fails_closed_without_crossing_handles(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    manifest = _write_manifest_files(tmp_path, days=2)
    target = manifest.workbooks[SectorTicker.XLC]
    replacement = tmp_path / "replacement.xlsx"
    _write_workbook(replacement, [(_START, 900), (_START + timedelta(days=1), 901)])
    real_load_workbook = subject._load_workbook
    swapped = False

    def swapping_load_workbook(source: object, **kwargs: object) -> object:
        nonlocal swapped
        if not swapped:
            target.unlink()
            replacement.replace(target)
            swapped = True
        return real_load_workbook(source, **kwargs)

    monkeypatch.setattr(subject, "_load_workbook", swapping_load_workbook)
    parsed = parse_private_nav_workbooks(manifest, repository_root=REPOSITORY_ROOT)

    assert swapped
    assert len(parsed.series_by_ticker) == 11
    assert _failure_for(parsed, SectorTicker.XLC).issue_code is WorkbookIssueCode.OPEN


def test_in_place_write_after_preflight_cannot_change_openpyxl_snapshot(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    manifest = _write_manifest_files(tmp_path, days=2)
    target = manifest.workbooks[SectorTicker.XLC]
    original_bytes = target.read_bytes()
    replacement = tmp_path / "in-place-replacement.xlsx"
    _write_workbook(
        replacement,
        [(_START, 700), (_START + timedelta(days=1), 701), (_START + timedelta(days=2), 702)],
    )
    replacement_bytes = replacement.read_bytes()
    real_load_workbook = subject._load_workbook
    mutated = False

    def mutating_load_workbook(source: Any, **kwargs: Any) -> Any:
        nonlocal mutated
        if not mutated:
            source.seek(0)
            assert source.read() == original_bytes
            source.seek(0)
            target.write_bytes(replacement_bytes)
            mutated = True
        return real_load_workbook(source, **kwargs)

    monkeypatch.setattr(subject, "_load_workbook", mutating_load_workbook)
    parsed = parse_private_nav_workbooks(manifest, repository_root=REPOSITORY_ROOT)

    assert mutated
    assert len(parsed.series_by_ticker) == 11
    assert _failure_for(parsed, SectorTicker.XLC).issue_code is WorkbookIssueCode.OPEN


def test_workbooks_are_opened_and_closed_strictly_sequentially(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    invalid_rows = _rows(6)
    invalid_rows[2] = (invalid_rows[2][0], False)
    manifest = _write_manifest_files(
        tmp_path,
        days=6,
        rows_by_ticker={SectorTicker.XLC: invalid_rows},
    )
    real_load_workbook = subject._load_workbook
    active = 0
    max_active = 0
    seen_options: list[dict[str, Any]] = []

    class TrackingWorkbook:
        def __init__(self, wrapped: Any) -> None:
            self._wrapped = wrapped

        @property
        def worksheets(self) -> Any:
            return self._wrapped.worksheets

        def close(self) -> None:
            nonlocal active
            try:
                self._wrapped.close()
            finally:
                active -= 1

    def tracking_load_workbook(*args: Any, **kwargs: Any) -> TrackingWorkbook:
        nonlocal active, max_active
        seen_options.append(dict(kwargs))
        wrapped = real_load_workbook(*args, **kwargs)
        active += 1
        max_active = max(max_active, active)
        return TrackingWorkbook(wrapped)

    monkeypatch.setattr(subject, "_load_workbook", tracking_load_workbook)
    parsed = parse_private_nav_workbooks(manifest, repository_root=REPOSITORY_ROOT)

    assert len(parsed.series_by_ticker) == 11
    assert len(parsed.failures) == 1
    assert active == 0
    assert max_active == 1
    assert (
        seen_options
        == [
            {
                "read_only": True,
                "data_only": True,
                "keep_links": False,
                "keep_vba": False,
            }
        ]
        * 12
    )


def test_private_input_module_has_no_network_or_work_component_imports() -> None:
    tree = ast.parse(Path(subject.__file__).read_text(encoding="utf-8"))
    imported_modules: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            imported_modules.update(alias.name for alias in node.names)
        elif isinstance(node, ast.ImportFrom) and node.module is not None:
            imported_modules.add(node.module)

    forbidden = (
        "httpx",
        "requests",
        "socket",
        "investo.sources",
        "investo.briefing",
        "investo.publisher",
        "investo.notifier",
    )
    assert not any(
        module == prefix or module.startswith(f"{prefix}.")
        for module in imported_modules
        for prefix in forbidden
    )
