"""Bounded local XLSX input boundary for u139 private NAV validation.

The adapter accepts only an explicit, fixed-universe manifest.  It performs all
repository/path checks and a bounded OOXML ZIP/XML preflight before ``openpyxl``
sees any workbook.  Provider bytes, paths, cell values, and exception messages do
not cross this module's redacted result boundary.

References: FR-022, FD R1-R14/R31-R33, NFR AC-1.1-AC-2.4/AC-3.1.
"""

from __future__ import annotations

import hashlib
import json
import os
import stat
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO
from itertools import pairwise
from pathlib import Path, PurePosixPath
from typing import BinaryIO, Final, Protocol, cast
from zipfile import ZipFile, ZipInfo

from defusedxml.ElementTree import iterparse
from openpyxl import load_workbook as _openpyxl_load_workbook  # type: ignore[import-untyped]
from pydantic import ValidationError

from investo.models import (
    ALL_SECTOR_TICKERS,
    BENCHMARK_TICKER,
    FIXED_SECTOR_UNIVERSE,
    SECTOR_TICKERS,
    CoverageSummary,
    DiagnosticIssueCode,
    NavPoint,
    NavSeries,
    ParsedWorkbookSet,
    PrivateDiagnostic,
    PrivateWorkbookManifest,
    SectorCoverageStatus,
    SectorSeriesBundle,
    SectorTicker,
    SectorUniverse,
    WorkbookFailure,
    WorkbookIssueCode,
)

MAX_MANIFEST_BYTES: Final[int] = 64 * 1024
MAX_COMPRESSED_WORKBOOK_BYTES: Final[int] = 8 * 1024 * 1024
MAX_UNCOMPRESSED_WORKBOOK_BYTES: Final[int] = 64 * 1024 * 1024
MAX_COMPRESSION_RATIO: Final[int] = 100
MAX_ZIP_MEMBERS: Final[int] = 2_000
MAX_WORKSHEETS: Final[int] = 20
MAX_WORKSHEET_CELLS: Final[int] = 250_000

_XLSX_CONTENT_TYPE: Final[str] = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
)
_XLSX_WORKSHEET_CONTENT_TYPE: Final[str] = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
)
_WORKSHEET_RELATIONSHIP_TYPES: Final[frozenset[str]] = frozenset(
    {
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet",
        "http://purl.oclc.org/ooxml/officeDocument/relationships/worksheet",
    }
)
_REQUIRED_PACKAGE_MEMBERS: Final[frozenset[str]] = frozenset(
    {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    }
)
_load_workbook = _openpyxl_load_workbook


class PrivateInputError(ValueError):
    """A run-level private-input rejection whose text is always redacted."""

    def __init__(self, issue_code: DiagnosticIssueCode) -> None:
        self.issue_code = issue_code
        super().__init__(issue_code.value)


class _XmlElement(Protocol):
    tag: str
    attrib: Mapping[str, str]

    def clear(self) -> None: ...


class _ReadOnlyWorksheet(Protocol):
    def reset_dimensions(self) -> None: ...

    def iter_rows(
        self,
        *,
        min_row: int | None = None,
        values_only: bool = False,
    ) -> Iterable[Sequence[object]]: ...


class _ReadOnlyWorkbook(Protocol):
    worksheets: Sequence[_ReadOnlyWorksheet]

    def close(self) -> None: ...


@dataclass(frozen=True, slots=True)
class _HeaderLocation:
    worksheet_index: int
    row_number: int
    date_column: int
    nav_column: int


@dataclass(frozen=True, slots=True)
class _FileIdentity:
    device: int
    inode: int
    size: int
    modified_ns: int
    changed_ns: int

    @classmethod
    def from_stat(cls, value: os.stat_result) -> _FileIdentity:
        return cls(
            device=value.st_dev,
            inode=value.st_ino,
            size=value.st_size,
            modified_ns=value.st_mtime_ns,
            changed_ns=value.st_ctime_ns,
        )


@dataclass(frozen=True, slots=True)
class _ResolvedWorkbook:
    path: Path
    identity: _FileIdentity


@dataclass(slots=True)
class _WorkbookRejectedError(Exception):
    issue_code: WorkbookIssueCode
    row_count: int | None = None
    first_date: date | None = None
    latest_date: date | None = None

    def as_failure(self, ticker: SectorTicker) -> WorkbookFailure:
        return WorkbookFailure(
            ticker=ticker,
            issue_code=self.issue_code,
            row_count=self.row_count,
            first_date=self.first_date,
            latest_date=self.latest_date,
        )


class _DuplicateJsonKeyError(ValueError):
    pass


_DIAGNOSTIC_BY_WORKBOOK_ISSUE: Final[dict[WorkbookIssueCode, DiagnosticIssueCode]] = {
    WorkbookIssueCode.OPEN: DiagnosticIssueCode.WORKBOOK_OPEN,
    WorkbookIssueCode.HEADER: DiagnosticIssueCode.WORKBOOK_HEADER,
    WorkbookIssueCode.DATE: DiagnosticIssueCode.WORKBOOK_DATE,
    WorkbookIssueCode.NAV: DiagnosticIssueCode.WORKBOOK_NAV,
    WorkbookIssueCode.ORDER: DiagnosticIssueCode.WORKBOOK_ORDER,
    WorkbookIssueCode.DUPLICATE_DATE: DiagnosticIssueCode.WORKBOOK_DUPLICATE_DATE,
}


def read_private_workbook_manifest(
    manifest_path: Path,
    *,
    repository_root: Path,
    expected_universe: SectorUniverse = FIXED_SECTOR_UNIVERSE,
) -> PrivateWorkbookManifest:
    """Read one explicit untracked JSON manifest and resolve its workbook paths.

    Every failure is reduced to a stable manifest issue code.  The original path,
    JSON value, and validation exception never enter the raised error.
    """

    if expected_universe != FIXED_SECTOR_UNIVERSE:
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_UNIVERSE)
    repository = _resolve_repository_root(repository_root)
    resolved_manifest, manifest_identity = _resolve_external_regular_file(
        manifest_path,
        repository_root=repository,
        expected_suffix=".json",
    )
    try:
        with _open_verified_binary(resolved_manifest, manifest_identity) as stream:
            manifest_bytes = stream.read(MAX_MANIFEST_BYTES + 1)
            if _FileIdentity.from_stat(
                os.fstat(stream.fileno())
            ) != manifest_identity or not _path_matches_identity(
                resolved_manifest, manifest_identity
            ):
                raise PrivateInputError(DiagnosticIssueCode.MANIFEST_PATH)
        if len(manifest_bytes) > MAX_MANIFEST_BYTES:
            raise PrivateInputError(DiagnosticIssueCode.MANIFEST_SCHEMA)
        payload = json.loads(manifest_bytes, object_pairs_hook=_reject_duplicate_json_keys)
    except PrivateInputError:
        raise
    except _DuplicateJsonKeyError:
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_SCHEMA) from None
    except OSError:
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_PATH) from None
    except (UnicodeDecodeError, json.JSONDecodeError):
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_SCHEMA) from None

    if (
        not isinstance(payload, dict)
        or set(payload) != {"schema_version", "workbooks"}
        or payload.get("schema_version") != 1
    ):
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_SCHEMA)
    workbooks = payload.get("workbooks")
    if not isinstance(workbooks, dict) or set(workbooks) != {
        ticker.value for ticker in ALL_SECTOR_TICKERS
    }:
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_UNIVERSE)

    try:
        manifest = PrivateWorkbookManifest.model_validate(payload)
    except ValidationError:
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_PATH) from None
    resolved, _ = _resolve_manifest_workbooks(manifest, repository_root=repository)
    return resolved


def parse_private_nav_workbooks(
    manifest: PrivateWorkbookManifest,
    *,
    repository_root: Path,
    expected_universe: SectorUniverse = FIXED_SECTOR_UNIVERSE,
) -> ParsedWorkbookSet:
    """Parse the fixed manifest into exactly one success or failure per ticker."""

    parsed, _ = _parse_manifest(
        manifest,
        repository_root=repository_root,
        expected_universe=expected_universe,
        include_fingerprint=False,
    )
    return parsed


def load_private_nav_workbooks(
    manifest: PrivateWorkbookManifest,
    *,
    repository_root: Path,
    expected_universe: SectorUniverse = FIXED_SECTOR_UNIVERSE,
) -> SectorSeriesBundle:
    """Load an explicit local manifest into one redacted canonical NAV bundle."""

    parsed, input_fingerprint = _parse_manifest(
        manifest,
        repository_root=repository_root,
        expected_universe=expected_universe,
        include_fingerprint=True,
    )
    return _build_bundle(parsed, input_fingerprint=input_fingerprint)


def _parse_manifest(
    manifest: PrivateWorkbookManifest,
    *,
    repository_root: Path,
    expected_universe: SectorUniverse,
    include_fingerprint: bool,
) -> tuple[ParsedWorkbookSet, str | None]:
    if expected_universe != FIXED_SECTOR_UNIVERSE:
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_UNIVERSE)
    repository = _resolve_repository_root(repository_root)
    _, resolved_workbooks = _resolve_manifest_workbooks(manifest, repository_root=repository)

    series_by_ticker: dict[SectorTicker, NavSeries] = {}
    failures: list[WorkbookFailure] = []
    digests: dict[SectorTicker, bytes] = {}
    for ticker in ALL_SECTOR_TICKERS:
        result, workbook_digest = _parse_verified_workbook(ticker, resolved_workbooks[ticker])
        if isinstance(result, WorkbookFailure):
            failures.append(result)
        else:
            series_by_ticker[ticker] = result
        if workbook_digest is not None:
            digests[ticker] = workbook_digest

    input_fingerprint = (
        _fingerprint_workbook_digests(digests)
        if include_fingerprint and len(digests) == len(ALL_SECTOR_TICKERS)
        else None
    )

    return (
        ParsedWorkbookSet(series_by_ticker=series_by_ticker, failures=tuple(failures)),
        input_fingerprint,
    )


def _resolve_repository_root(repository_root: Path) -> Path:
    try:
        resolved = repository_root.resolve(strict=True)
    except OSError:
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_PATH) from None
    if not resolved.is_dir():
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_PATH)
    return resolved


def _resolve_manifest_workbooks(
    manifest: PrivateWorkbookManifest,
    *,
    repository_root: Path,
) -> tuple[PrivateWorkbookManifest, Mapping[SectorTicker, _ResolvedWorkbook]]:
    resolved: dict[SectorTicker, Path] = {}
    resolved_workbooks: dict[SectorTicker, _ResolvedWorkbook] = {}
    identities: set[tuple[int, int]] = set()
    for ticker in ALL_SECTOR_TICKERS:
        path, file_identity = _resolve_external_regular_file(
            manifest.workbooks[ticker],
            repository_root=repository_root,
            expected_suffix=".xlsx",
        )
        identity = (file_identity.device, file_identity.inode)
        if identity in identities:
            raise PrivateInputError(DiagnosticIssueCode.MANIFEST_PATH)
        identities.add(identity)
        resolved[ticker] = path
        resolved_workbooks[ticker] = _ResolvedWorkbook(path=path, identity=file_identity)
    return (
        PrivateWorkbookManifest(schema_version=1, workbooks=resolved),
        resolved_workbooks,
    )


def _resolve_external_regular_file(
    path: Path,
    *,
    repository_root: Path,
    expected_suffix: str,
) -> tuple[Path, _FileIdentity]:
    if not path.is_absolute() or path.suffix.lower() != expected_suffix:
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_PATH)
    try:
        original_stat = path.lstat()
        resolved = path.resolve(strict=True)
        resolved_stat = resolved.stat()
    except OSError:
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_PATH) from None
    if stat.S_ISLNK(original_stat.st_mode) or not stat.S_ISREG(original_stat.st_mode):
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_PATH)
    if not stat.S_ISREG(resolved_stat.st_mode) or resolved.suffix.lower() != expected_suffix:
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_PATH)
    if _is_within(resolved, repository_root):
        raise PrivateInputError(DiagnosticIssueCode.MANIFEST_PATH)
    return resolved, _FileIdentity.from_stat(resolved_stat)


def _is_within(path: Path, root: Path) -> bool:
    return path == root or root in path.parents


def _reject_duplicate_json_keys(pairs: list[tuple[str, object]]) -> dict[str, object]:
    value: dict[str, object] = {}
    for key, item in pairs:
        if key in value:
            raise _DuplicateJsonKeyError
        value[key] = item
    return value


def _open_verified_binary(path: Path, expected: _FileIdentity) -> BinaryIO:
    no_follow = getattr(os, "O_NOFOLLOW", None)
    if no_follow is None:
        raise OSError("no-follow file open is unavailable")
    flags = os.O_RDONLY | no_follow | getattr(os, "O_CLOEXEC", 0)
    descriptor = os.open(path, flags)
    try:
        opened_stat = os.fstat(descriptor)
        if not stat.S_ISREG(opened_stat.st_mode):
            raise OSError("private input is not a regular file")
        if _FileIdentity.from_stat(opened_stat) != expected:
            raise OSError("private input identity changed")
        return os.fdopen(descriptor, "rb", closefd=True)
    except Exception:
        os.close(descriptor)
        raise


def _path_matches_identity(path: Path, expected: _FileIdentity) -> bool:
    try:
        path_stat = path.lstat()
    except OSError:
        return False
    return stat.S_ISREG(path_stat.st_mode) and _FileIdentity.from_stat(path_stat) == expected


def _preflight_xlsx(path: Path) -> None:
    try:
        with path.open("rb") as stream:
            _preflight_xlsx_stream(stream, compressed_size=os.fstat(stream.fileno()).st_size)
    except _WorkbookRejectedError:
        raise
    except Exception as exc:
        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN) from exc


def _preflight_xlsx_stream(stream: BinaryIO, *, compressed_size: int) -> None:
    try:
        if compressed_size > MAX_COMPRESSED_WORKBOOK_BYTES:
            raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
        stream.seek(0)
        with ZipFile(stream) as package:
            infos = package.infolist()
            _validate_zip_metadata(infos, compressed_size=compressed_size)
            infos_by_name = {info.filename: info for info in infos}
            if not _REQUIRED_PACKAGE_MEMBERS.issubset(infos_by_name):
                raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
            content_types = _read_content_types(
                package,
                infos_by_name["[Content_Types].xml"],
            )
            _validate_workbook_shape(package, infos_by_name, content_types=content_types)
            if package.testzip() is not None:
                raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
    except _WorkbookRejectedError:
        raise
    except Exception as exc:
        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN) from exc
    finally:
        stream.seek(0)


def _validate_zip_metadata(infos: Sequence[ZipInfo], *, compressed_size: int) -> None:
    if not infos or len(infos) > MAX_ZIP_MEMBERS:
        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
    names: set[str] = set()
    uncompressed_total = 0
    compressed_total = 0
    for info in infos:
        original_name = info.orig_filename
        normalized_name = original_name.replace("\\", "/")
        member_path = PurePosixPath(normalized_name)
        if (
            "\x00" in original_name
            or "\\" in original_name
            or normalized_name in names
            or member_path.is_absolute()
            or ".." in member_path.parts
            or (
                member_path.parts
                and len(member_path.parts[0]) >= 2
                and member_path.parts[0][1] == ":"
            )
            or info.flag_bits & 0x1
        ):
            raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
        names.add(normalized_name)
        uncompressed_total += info.file_size
        compressed_total += info.compress_size
        if info.file_size > MAX_UNCOMPRESSED_WORKBOOK_BYTES:
            raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
    if uncompressed_total > MAX_UNCOMPRESSED_WORKBOOK_BYTES:
        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
    if uncompressed_total and compressed_total == 0:
        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
    if compressed_total and uncompressed_total > compressed_total * MAX_COMPRESSION_RATIO:
        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
    if compressed_size <= 0:
        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)


def _read_content_types(package: ZipFile, info: ZipInfo) -> Mapping[str, str]:
    xlsx_main_type_found = False
    overrides: dict[str, str] = {}
    with package.open(info) as stream:
        for _, raw_element in iterparse(stream, events=("end",)):
            element = cast(_XmlElement, raw_element)
            if _local_name(element.tag) == "Override":
                part_name = element.attrib.get("PartName")
                content_type = element.attrib.get("ContentType")
                if (
                    part_name is None
                    or content_type is None
                    or part_name in overrides
                    or not part_name.startswith("/")
                ):
                    raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
                overrides[part_name] = content_type
                if part_name == "/xl/workbook.xml" and content_type == _XLSX_CONTENT_TYPE:
                    xlsx_main_type_found = True
            element.clear()
    if not xlsx_main_type_found:
        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
    return overrides


def _validate_workbook_shape(
    package: ZipFile,
    infos_by_name: Mapping[str, ZipInfo],
    *,
    content_types: Mapping[str, str],
) -> None:
    sheet_relationship_ids = _read_sheet_relationship_ids(
        package,
        infos_by_name["xl/workbook.xml"],
    )
    worksheet_targets = _read_worksheet_targets(
        package,
        infos_by_name["xl/_rels/workbook.xml.rels"],
        sheet_relationship_ids=sheet_relationship_ids,
    )
    if not worksheet_targets or len(worksheet_targets) > MAX_WORKSHEETS:
        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
    worksheet_infos: list[ZipInfo] = []
    for target in worksheet_targets:
        info = infos_by_name.get(target)
        if info is None or content_types.get(f"/{target}") != _XLSX_WORKSHEET_CONTENT_TYPE:
            raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
        worksheet_infos.append(info)

    cell_count = 0
    for info in worksheet_infos:
        with package.open(info) as stream:
            for _, raw_element in iterparse(stream, events=("end",)):
                element = cast(_XmlElement, raw_element)
                if _local_name(element.tag) == "c":
                    cell_count += 1
                    if cell_count > MAX_WORKSHEET_CELLS:
                        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
                element.clear()


def _read_sheet_relationship_ids(package: ZipFile, info: ZipInfo) -> tuple[str, ...]:
    relationship_ids: list[str] = []
    with package.open(info) as stream:
        for _, raw_element in iterparse(stream, events=("end",)):
            element = cast(_XmlElement, raw_element)
            if _local_name(element.tag) == "sheet":
                relationship_id = _attribute_by_local_name(element.attrib, "id")
                if relationship_id is None or relationship_id in relationship_ids:
                    raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
                relationship_ids.append(relationship_id)
                if len(relationship_ids) > MAX_WORKSHEETS:
                    raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
            element.clear()
    if not relationship_ids:
        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
    return tuple(relationship_ids)


def _read_worksheet_targets(
    package: ZipFile,
    info: ZipInfo,
    *,
    sheet_relationship_ids: Sequence[str],
) -> tuple[str, ...]:
    relationships: dict[str, tuple[str, str]] = {}
    with package.open(info) as stream:
        for _, raw_element in iterparse(stream, events=("end",)):
            element = cast(_XmlElement, raw_element)
            if _local_name(element.tag) == "Relationship":
                relationship_id = element.attrib.get("Id")
                relationship_type = element.attrib.get("Type")
                target = element.attrib.get("Target")
                if (
                    relationship_id is None
                    or relationship_type is None
                    or target is None
                    or relationship_id in relationships
                    or element.attrib.get("TargetMode") is not None
                ):
                    raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
                relationships[relationship_id] = (relationship_type, target)
            element.clear()

    targets: list[str] = []
    for relationship_id in sheet_relationship_ids:
        relationship = relationships.get(relationship_id)
        if relationship is None:
            raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
        relationship_type, target = relationship
        if relationship_type not in _WORKSHEET_RELATIONSHIP_TYPES:
            raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
        normalized_target = _normalize_relationship_target(target)
        if normalized_target in targets:
            raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
        targets.append(normalized_target)
    return tuple(targets)


def _normalize_relationship_target(target: str) -> str:
    if "\x00" in target or "\\" in target:
        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
    target_path = PurePosixPath(target)
    if (
        ".." in target_path.parts
        or not target_path.parts
        or (
            target_path.parts[0] != "/"
            and len(target_path.parts[0]) >= 2
            and target_path.parts[0][1] == ":"
        )
    ):
        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
    if target_path.is_absolute():
        normalized = PurePosixPath(*target_path.parts[1:])
    else:
        normalized = PurePosixPath("xl", *target_path.parts)
    if not normalized.parts or normalized.parts[0] != "xl":
        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
    return normalized.as_posix()


def _attribute_by_local_name(attributes: Mapping[str, str], name: str) -> str | None:
    matches = tuple(value for key, value in attributes.items() if _local_name(key) == name)
    if len(matches) != 1:
        return None
    return matches[0]


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _read_verified_workbook_bytes(
    stream: BinaryIO,
    resolved: _ResolvedWorkbook,
) -> bytes:
    stream.seek(0)
    workbook_bytes = stream.read(MAX_COMPRESSED_WORKBOOK_BYTES + 1)
    if (
        len(workbook_bytes) > MAX_COMPRESSED_WORKBOOK_BYTES
        or _FileIdentity.from_stat(os.fstat(stream.fileno())) != resolved.identity
        or not _path_matches_identity(resolved.path, resolved.identity)
    ):
        raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
    return workbook_bytes


def _fingerprint_workbook_digests(digests: Mapping[SectorTicker, bytes]) -> str:
    digest = hashlib.sha256(b"investo-u139-private-input-v1\x00")
    for ticker in ALL_SECTOR_TICKERS:
        ticker_bytes = ticker.value.encode("ascii")
        workbook_digest = digests[ticker]
        if len(workbook_digest) != hashlib.sha256().digest_size:
            raise ValueError("workbook digest must be SHA-256")
        digest.update(len(ticker_bytes).to_bytes(1, "big"))
        digest.update(ticker_bytes)
        digest.update(workbook_digest)
    return f"sha256:{digest.hexdigest()}"


def _parse_verified_workbook(
    ticker: SectorTicker,
    resolved: _ResolvedWorkbook,
) -> tuple[NavSeries | WorkbookFailure, bytes | None]:
    workbook: _ReadOnlyWorkbook | None = None
    result: NavSeries | WorkbookFailure = WorkbookFailure(
        ticker=ticker,
        issue_code=WorkbookIssueCode.OPEN,
    )
    workbook_digest: bytes | None = None
    close_failed = False
    try:
        with _open_verified_binary(resolved.path, resolved.identity) as stream:
            workbook_bytes = _read_verified_workbook_bytes(stream, resolved)
            workbook_digest = hashlib.sha256(workbook_bytes).digest()
            with BytesIO(workbook_bytes) as snapshot:
                _preflight_xlsx_stream(snapshot, compressed_size=len(workbook_bytes))
                try:
                    workbook = cast(
                        _ReadOnlyWorkbook,
                        _load_workbook(
                            snapshot,
                            read_only=True,
                            data_only=True,
                            keep_links=False,
                            keep_vba=False,
                        ),
                    )
                    result = _parse_open_workbook(ticker, workbook)
                except _WorkbookRejectedError as rejected:
                    result = rejected.as_failure(ticker)
                except Exception:
                    result = WorkbookFailure(ticker=ticker, issue_code=WorkbookIssueCode.OPEN)
                finally:
                    if workbook is not None:
                        try:
                            workbook.close()
                        except Exception:
                            close_failed = True

            current_identity = _FileIdentity.from_stat(os.fstat(stream.fileno()))
            if current_identity != resolved.identity or not _path_matches_identity(
                resolved.path, resolved.identity
            ):
                raise _WorkbookRejectedError(WorkbookIssueCode.OPEN)
    except _WorkbookRejectedError as rejected:
        result = rejected.as_failure(ticker)
    except Exception:
        result = WorkbookFailure(ticker=ticker, issue_code=WorkbookIssueCode.OPEN)
    if close_failed:
        result = WorkbookFailure(ticker=ticker, issue_code=WorkbookIssueCode.OPEN)
    return result, workbook_digest


def _parse_open_workbook(ticker: SectorTicker, workbook: _ReadOnlyWorkbook) -> NavSeries:
    headers: list[_HeaderLocation] = []
    for worksheet_index, worksheet in enumerate(workbook.worksheets):
        worksheet.reset_dimensions()
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            columns = _header_columns(row)
            if columns is not None:
                headers.append(
                    _HeaderLocation(
                        worksheet_index=worksheet_index,
                        row_number=row_number,
                        date_column=columns[0],
                        nav_column=columns[1],
                    )
                )
                break
    if len(headers) != 1:
        raise _WorkbookRejectedError(WorkbookIssueCode.HEADER)

    header = headers[0]
    worksheet = workbook.worksheets[header.worksheet_index]
    points: list[NavPoint] = []
    for row in worksheet.iter_rows(min_row=header.row_number + 1, values_only=True):
        if _header_columns(row) is not None:
            raise _rejection_with_range(WorkbookIssueCode.HEADER, points)
        date_value = _column_value(row, header.date_column)
        nav_value = _column_value(row, header.nav_column)
        if _is_blank(date_value) and _is_blank(nav_value):
            continue
        try:
            trading_date = _coerce_trading_date(date_value)
        except ValueError as exc:
            raise _rejection_with_range(WorkbookIssueCode.DATE, points) from exc
        try:
            nav = _coerce_nav(nav_value)
            point = NavPoint(trading_date=trading_date, nav=nav)
        except (InvalidOperation, ValidationError, ValueError) as exc:
            raise _rejection_with_range(WorkbookIssueCode.NAV, points) from exc
        points.append(point)

    if len(points) < 2:
        raise _rejection_with_range(WorkbookIssueCode.DATE, points)
    dates = tuple(point.trading_date for point in points)
    if len(set(dates)) != len(dates):
        raise _rejection_with_range(WorkbookIssueCode.DUPLICATE_DATE, points)
    ascending = all(current < following for current, following in pairwise(dates))
    descending = all(current > following for current, following in pairwise(dates))
    if not ascending and not descending:
        raise _rejection_with_range(WorkbookIssueCode.ORDER, points)
    if descending:
        points.reverse()
    return NavSeries(
        ticker=ticker,
        points=tuple(points),
        first_date=points[0].trading_date,
        latest_date=points[-1].trading_date,
    )


def _header_columns(row: Sequence[object]) -> tuple[int, int] | None:
    normalized = tuple(_normalize_header(value) for value in row)
    date_columns = tuple(index for index, value in enumerate(normalized) if value == "date")
    nav_columns = tuple(index for index, value in enumerate(normalized) if value == "nav")
    if not date_columns or not nav_columns:
        return None
    if len(date_columns) != 1 or len(nav_columns) != 1:
        raise _WorkbookRejectedError(WorkbookIssueCode.HEADER)
    return date_columns[0], nav_columns[0]


def _normalize_header(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    return " ".join(value.strip().split()).casefold()


def _column_value(row: Sequence[object], index: int) -> object:
    return row[index] if index < len(row) else None


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _coerce_trading_date(value: object) -> date:
    if isinstance(value, datetime):
        if value.tzinfo is not None or value.time() != time.min:
            raise ValueError("invalid date")
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        candidate = value.strip()
        parsed = date.fromisoformat(candidate)
        if candidate != parsed.isoformat():
            raise ValueError("invalid date")
        return parsed
    raise ValueError("invalid date")


def _coerce_nav(value: object) -> Decimal:
    if isinstance(value, bool) or value is None:
        raise ValueError("invalid NAV")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(str(value))
    if isinstance(value, float):
        return Decimal(repr(value))
    if isinstance(value, str):
        candidate = value.strip()
        if not candidate:
            raise ValueError("invalid NAV")
        return Decimal(candidate)
    raise ValueError("invalid NAV")


def _rejection_with_range(
    issue_code: WorkbookIssueCode,
    points: Sequence[NavPoint],
) -> _WorkbookRejectedError:
    dates = tuple(point.trading_date for point in points)
    return _WorkbookRejectedError(
        issue_code=issue_code,
        row_count=len(points),
        first_date=min(dates) if dates else None,
        latest_date=max(dates) if dates else None,
    )


def _build_bundle(
    parsed: ParsedWorkbookSet,
    *,
    input_fingerprint: str | None,
) -> SectorSeriesBundle:
    series_by_ticker = parsed.series_by_ticker
    benchmark = series_by_ticker.get(BENCHMARK_TICKER)
    sectors = tuple(
        series_by_ticker[ticker] for ticker in SECTOR_TICKERS if ticker in series_by_ticker
    )
    missing_tickers = tuple(ticker for ticker in SECTOR_TICKERS if ticker not in series_by_ticker)
    diagnostics = [
        PrivateDiagnostic(
            issue_code=_DIAGNOSTIC_BY_WORKBOOK_ISSUE[failure.issue_code],
            ticker=failure.ticker,
            row_count=failure.row_count,
            first_date=failure.first_date,
            latest_date=failure.latest_date,
        )
        for failure in parsed.failures
    ]
    reason_codes = [diagnostic.issue_code for diagnostic in diagnostics]

    mismatch = benchmark is not None and any(
        series.latest_date != benchmark.latest_date for series in sectors
    )
    if benchmark is None:
        status = SectorCoverageStatus.INSUFFICIENT
        common_as_of = None
        benchmark_observation_count = 0
        _append_bundle_reason(
            DiagnosticIssueCode.BUNDLE_SPY_MISSING,
            diagnostics=diagnostics,
            reason_codes=reason_codes,
        )
        _append_bundle_reason(
            DiagnosticIssueCode.BUNDLE_COVERAGE_INSUFFICIENT,
            diagnostics=diagnostics,
            reason_codes=reason_codes,
        )
    elif mismatch:
        status = SectorCoverageStatus.INSUFFICIENT
        common_as_of = None
        benchmark_observation_count = len(benchmark.points)
        _append_bundle_reason(
            DiagnosticIssueCode.BUNDLE_AS_OF_MISMATCH,
            diagnostics=diagnostics,
            reason_codes=reason_codes,
        )
    else:
        common_as_of = benchmark.latest_date
        benchmark_observation_count = len(benchmark.points)
        available_sector_count = len(sectors)
        if available_sector_count < 8 or benchmark_observation_count < 6:
            status = SectorCoverageStatus.INSUFFICIENT
            _append_bundle_reason(
                DiagnosticIssueCode.BUNDLE_COVERAGE_INSUFFICIENT,
                diagnostics=diagnostics,
                reason_codes=reason_codes,
            )
        elif benchmark_observation_count <= 21:
            status = SectorCoverageStatus.WARMING_UP
        elif available_sector_count <= 10:
            status = SectorCoverageStatus.PARTIAL
        else:
            status = SectorCoverageStatus.NORMAL

    coverage = CoverageSummary(
        status=status,
        available_sector_count=len(sectors),
        benchmark_available=benchmark is not None,
        common_as_of=common_as_of,
        benchmark_observation_count=benchmark_observation_count,
        missing_tickers=missing_tickers,
        reason_codes=tuple(reason_codes),
    )
    return SectorSeriesBundle(
        as_of_date=common_as_of,
        benchmark=benchmark,
        sectors=sectors,
        coverage=coverage,
        diagnostics=tuple(diagnostics),
        input_fingerprint=input_fingerprint,
    )


def _append_bundle_reason(
    issue_code: DiagnosticIssueCode,
    *,
    diagnostics: list[PrivateDiagnostic],
    reason_codes: list[DiagnosticIssueCode],
) -> None:
    diagnostics.append(PrivateDiagnostic(issue_code=issue_code))
    reason_codes.append(issue_code)


__all__ = [
    "PrivateInputError",
    "load_private_nav_workbooks",
    "parse_private_nav_workbooks",
    "read_private_workbook_manifest",
]
